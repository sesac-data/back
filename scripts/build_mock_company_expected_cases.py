import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
SOURCE_ROOT = ROOT_DIR / "data" / "가상기업데이터_20개"
FIXTURE_DIR = ROOT_DIR / "tests" / "fixtures" / "mock_company_cases"
OUTPUT_PATH = (
    ROOT_DIR
    / "output"
    / "mock_company_case_checks"
    / "mock_company_expected_case_inventory.json"
)
PLAN_PATH = ROOT_DIR / "tasks" / "mock_company_test_plan.md"

REQUIRED_FILES = {
    "company_features": "기업특징.txt",
    "business_registration": "사업자등록증.pdf",
    "employees": "직원명단.xlsx",
}

EXISTING_HANA_FOLDER = "04_주식회사_하나기계"
EXISTING_HANA_CASE_ID = "04_hana_machine"

POLICY_IDS = {
    "parental": "parental_leave_reduction_support",
    "replacement": "replacement_workshare_support",
}

MISSING_FIELD_CATALOG = {
    POLICY_IDS["parental"]: [
        "company.is_priority_support_enterprise",
        "leave_event.duration_days",
        "employee.child_age",
        "employee.child_age_months",
        "employee.child_school_grade",
    ],
    POLICY_IDS["replacement"]: [
        "company.is_priority_support_enterprise",
        "leave_event.duration_days",
        "replacement_worker.hire_date",
        "replacement_worker.employment_duration_days",
        "replacement_worker.is_new_hire",
    ],
}

RULE_INPUT_EXPANSION_BY_POLICY = {
    POLICY_IDS["parental"]: [
        "company.is_priority_support_enterprise",
        "leave_event.duration_days",
    ],
    POLICY_IDS["replacement"]: [],
}


def read_text(path):
    return path.read_text(encoding="utf-8")


def normalize_cell(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    return value


def parse_int(value):
    if value is None:
        return None
    match = re.search(r"\d+", str(value).replace(",", ""))
    if not match:
        return None
    return int(match.group(0))


def parse_features(text):
    parsed = {
        "data_type_code": None,
        "data_type_label": None,
        "target_employee_name": None,
        "target_employee_source_role": None,
        "program_type_source": None,
        "leave_start_date": None,
        "leave_end_date": None,
        "leave_duration_days": None,
        "replacement_worker_summary": None,
        "judgment_reason_text": "",
        "judgment_basis_text": "",
    }

    data_type_match = re.search(r"\[데이터 유형\]\s*(유형\d+)\s*-\s*(.+)", text)
    if data_type_match:
        parsed["data_type_code"] = data_type_match.group(1).strip()
        parsed["data_type_label"] = data_type_match.group(2).strip()

    employee_match = re.search(r"-\s*대상 근로자\s*:\s*([^\n(]+)(?:\(([^)]*)\))?", text)
    if employee_match:
        parsed["target_employee_name"] = employee_match.group(1).strip()
        parsed["target_employee_source_role"] = (
            employee_match.group(2).strip() if employee_match.group(2) else None
        )

    program_match = re.search(r"-\s*제도 유형\s*:\s*([^\n]+)", text)
    if program_match:
        parsed["program_type_source"] = program_match.group(1).strip()

    period_match = re.search(
        r"-\s*기간\s*:\s*(\d{4}-\d{2}-\d{2})\s*~\s*(\d{4}-\d{2}-\d{2})\s*\(총\s*(\d+)일\)",
        text,
    )
    if period_match:
        parsed["leave_start_date"] = period_match.group(1)
        parsed["leave_end_date"] = period_match.group(2)
        parsed["leave_duration_days"] = int(period_match.group(3))

    replacement_match = re.search(r"-\s*대체인력\s*:\s*([^\n]+(?:\n\s+[^\n]+)?)", text)
    if replacement_match:
        parsed["replacement_worker_summary"] = " ".join(
            replacement_match.group(1).split()
        )

    reason_match = re.search(
        r"\[지원금 수혜 판단 사유\](.*?)(?:\[판정 기준 근거\]|\Z)",
        text,
        flags=re.S,
    )
    if reason_match:
        parsed["judgment_reason_text"] = reason_match.group(1).strip()

    basis_match = re.search(r"\[판정 기준 근거\](.*)", text, flags=re.S)
    if basis_match:
        parsed["judgment_basis_text"] = basis_match.group(1).strip()

    return parsed


def parse_employee_workbook(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active

    company = {
        "company_name": normalize_cell(sheet["B1"].value),
        "business_registration_number": normalize_cell(sheet["B2"].value),
        "industry_code": str(normalize_cell(sheet["B3"].value) or ""),
        "insured_employee_count": parse_int(sheet["B4"].value),
    }

    headers = {}
    for cell in sheet[6]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    employees = []
    for row in range(7, sheet.max_row + 1):
        record = {}
        has_value = False
        for header, column in headers.items():
            value = normalize_cell(sheet.cell(row=row, column=column).value)
            if value is not None:
                has_value = True
            record[header] = value
        if has_value:
            employees.append(record)

    return {
        "company": company,
        "employees": employees,
    }


def find_target_employee(employees, features):
    target_name = features.get("target_employee_name")
    if target_name:
        for employee in employees:
            if employee.get("성명") == target_name:
                return employee

    for employee in employees:
        leave_type = str(employee.get("휴직유형") or "").strip()
        if leave_type and leave_type not in {"없음", "대체근무"}:
            return employee

    return None


def find_replacement_worker(employees):
    for employee in employees:
        leave_type = str(employee.get("휴직유형") or "").strip()
        new_hire = str(employee.get("신규채용") or "").strip().upper()
        if leave_type == "대체근무" or new_hire == "Y":
            return employee
    return None


def map_leave_type(source_value):
    value = str(source_value or "")
    if "근로시간" in value or "단축" in value:
        return "working_hour_reduction"
    if "육아휴직" in value:
        return "parental_leave"
    return "unknown"


def determine_target_policy(features):
    label = features.get("data_type_label") or ""
    program = features.get("program_type_source") or ""
    combined = f"{label} {program}"
    if "대체인력" in combined:
        return POLICY_IDS["replacement"]
    if "육아휴직" in combined or "근로시간 단축" in combined:
        return POLICY_IDS["parental"]
    return None


def determine_expected_status(features):
    label = features.get("data_type_label") or ""
    if "비대상" in label:
        return "ineligible"
    if "신청 대상" in label:
        return "eligible"
    return "unknown"


def determine_reason_code(policy_id, expected_status, features):
    text = " ".join(
        [
            features.get("data_type_label") or "",
            features.get("program_type_source") or "",
            features.get("judgment_reason_text") or "",
            features.get("judgment_basis_text") or "",
        ]
    )

    if policy_id == POLICY_IDS["parental"]:
        if expected_status == "eligible":
            return "parental_leave_or_reduction_duration_30_days_met"
        if expected_status == "ineligible":
            return "leave_duration_under_30_days"
        if "30일" in text and ("미만" in text or "부족" in text):
            return "leave_duration_under_30_days"
        return "parental_leave_reduction_requirement_not_met"

    if policy_id == POLICY_IDS["replacement"]:
        if expected_status == "eligible":
            return "replacement_worker_requirements_met"
        if "대체인력" in text and ("없" in text or "신규 채용" in text):
            if "2개월" not in text and "30일" not in text:
                return "replacement_worker_missing"
        if "2개월" in text and (
            "이르" in text or "빠르" in text or "이전" in text
        ):
            return "replacement_worker_hired_before_allowed_window"
        if "30일" in text and ("미만" in text or "부족" in text):
            return "replacement_worker_employment_duration_under_30_days"
        if "대체인력" in text and "없" in text:
            return "replacement_worker_missing"
        return "replacement_worker_requirement_not_met"

    return "manual_review_required"


def subtract_months(value, months):
    year = value.year
    month = value.month - months
    while month <= 0:
        year -= 1
        month += 12

    last_day_by_month = {
        1: 31,
        2: 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28,
        3: 31,
        4: 30,
        5: 31,
        6: 30,
        7: 31,
        8: 31,
        9: 30,
        10: 31,
        11: 30,
        12: 31,
    }
    day = min(value.day, last_day_by_month[month])
    return date(year, month, day)


def refine_replacement_reason_code(
    expected_status,
    leave_start_date,
    replacement_worker,
    replacement_duration,
):
    if expected_status == "eligible":
        return "replacement_worker_requirements_met"

    if not replacement_worker:
        return "replacement_worker_missing"

    hire_date = normalize_cell(replacement_worker.get("입사일"))
    if hire_date and leave_start_date:
        hire = datetime.strptime(hire_date, "%Y-%m-%d").date()
        leave_start = datetime.strptime(leave_start_date, "%Y-%m-%d").date()
        allowed_earliest = subtract_months(leave_start, 2)
        if hire < allowed_earliest:
            return "replacement_worker_hired_before_allowed_window"

    if replacement_duration is not None and replacement_duration < 30:
        return "replacement_worker_employment_duration_under_30_days"

    return "replacement_worker_requirement_not_met"


def requested_months(start_date, end_date):
    if not start_date or not end_date:
        return None
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()
    months = (end.year - start.year) * 12 + end.month - start.month + 1
    return max(months, 1)


def build_policy_mapping(target_policy_id, expected_status, reason_code):
    policies = [
        "parental_leave_reduction_support",
        "replacement_workshare_support",
        "worklife_balance_45_support",
        "childcare_flexible_start_support",
        "working_hours_reduction_support",
        "flexible_work_incent",
        "flexible_work_system_support",
    ]
    mapping = []
    for policy_id in policies:
        if policy_id == target_policy_id:
            mapping.append(
                {
                    "policy_id": policy_id,
                    "case_relevance": "target_policy",
                    "verification_status": "expected_result_draft",
                    "expected_status": expected_status,
                    "expected_reason_code": reason_code,
                }
            )
        else:
            mapping.append(
                {
                    "policy_id": policy_id,
                    "case_relevance": "not_primary_target",
                    "verification_status": "not_evaluated_in_this_draft",
                }
            )
    return mapping


def make_case_id(index, folder_name):
    if folder_name == EXISTING_HANA_FOLDER:
        return EXISTING_HANA_CASE_ID
    return f"mock_company_{index:02d}"


def build_case(folder, index):
    file_paths = {
        key: folder / filename for key, filename in REQUIRED_FILES.items()
    }
    file_presence = {
        key: path.exists() for key, path in file_paths.items()
    }
    missing_files = [
        str(path.name) for key, path in file_paths.items() if not file_presence[key]
    ]

    errors = []
    features = {}
    workbook_data = {"company": {}, "employees": []}

    if file_presence["company_features"]:
        features = parse_features(read_text(file_paths["company_features"]))
    else:
        errors.append("missing_company_features_file")

    if file_presence["employees"]:
        workbook_data = parse_employee_workbook(file_paths["employees"])
    else:
        errors.append("missing_employees_file")

    target_employee = find_target_employee(workbook_data["employees"], features) or {}
    replacement_worker = find_replacement_worker(workbook_data["employees"]) or {}
    target_policy_id = determine_target_policy(features)
    expected_status = determine_expected_status(features)
    leave_type = map_leave_type(
        target_employee.get("휴직유형") or features.get("program_type_source")
    )

    start_date = (
        features.get("leave_start_date")
        or normalize_cell(target_employee.get("시작일"))
    )
    end_date = (
        features.get("leave_end_date")
        or normalize_cell(target_employee.get("종료일"))
    )
    duration_days = features.get("leave_duration_days")

    if not target_policy_id:
        errors.append("target_policy_not_detected")
    if expected_status == "unknown":
        errors.append("expected_status_not_detected")
    if not target_employee:
        errors.append("target_employee_not_detected")

    company = workbook_data["company"]
    has_replacement_worker = bool(replacement_worker)
    replacement_duration = None
    if replacement_worker:
        replacement_start = normalize_cell(replacement_worker.get("시작일"))
        replacement_end = normalize_cell(replacement_worker.get("종료일"))
        if replacement_start and replacement_end:
            start = datetime.strptime(replacement_start, "%Y-%m-%d").date()
            end = datetime.strptime(replacement_end, "%Y-%m-%d").date()
            replacement_duration = (end - start).days + 1

    if target_policy_id == POLICY_IDS["replacement"]:
        reason_code = refine_replacement_reason_code(
            expected_status,
            start_date,
            replacement_worker,
            replacement_duration,
        )
    else:
        reason_code = determine_reason_code(
            target_policy_id,
            expected_status,
            features,
        )

    missing_fields = []
    for field in MISSING_FIELD_CATALOG.get(target_policy_id, []):
        if field.startswith("replacement_worker.") and not has_replacement_worker:
            missing_fields.append(field)
        elif field in {
            "employee.child_age",
            "employee.child_age_months",
            "employee.child_school_grade",
        }:
            missing_fields.append(field)
        elif field == "company.is_priority_support_enterprise":
            missing_fields.append(field)
        elif field == "leave_event.duration_days" and duration_days is None:
            missing_fields.append(field)

    comparable_likelihood = "medium"
    if target_policy_id == POLICY_IDS["replacement"]:
        comparable_likelihood = "high" if has_replacement_worker else "medium"
    if errors:
        comparable_likelihood = "low"

    case_id = make_case_id(index, folder.name)
    fixture_filename = f"{case_id}_expected.json"
    fixture_path = FIXTURE_DIR / fixture_filename
    if folder.name == EXISTING_HANA_FOLDER:
        fixture_path = FIXTURE_DIR / "04_hana_machine_expected.json"

    extracted_fields = {
        "company": {
            "company_name": company.get("company_name"),
            "business_registration_number": company.get(
                "business_registration_number"
            ),
            "industry_code": company.get("industry_code"),
            "insured_employee_count": company.get("insured_employee_count"),
            "is_priority_support_enterprise": None,
            "is_small_or_medium_company": True,
        },
        "employee": {
            "name": target_employee.get("성명"),
            "department_role": target_employee.get("부서/직무"),
            "employment_type_source_value": target_employee.get("고용형태"),
            "hire_date": normalize_cell(target_employee.get("입사일")),
            "leave_type": leave_type,
            "leave_type_source_value": target_employee.get("휴직유형"),
            "salary_monthly": target_employee.get("급여(원)"),
        },
        "leave_event": {
            "leave_type": leave_type,
            "start_date": start_date,
            "end_date": end_date,
            "duration_days": duration_days,
            "requested_months_for_current_adapter": requested_months(
                start_date,
                end_date,
            ),
            "has_replacement_worker": has_replacement_worker,
        },
        "replacement_worker": {
            "name": replacement_worker.get("성명"),
            "department_role": replacement_worker.get("부서/직무"),
            "employment_type_source_value": replacement_worker.get("고용형태"),
            "hire_date": normalize_cell(replacement_worker.get("입사일")),
            "work_start_date": normalize_cell(replacement_worker.get("시작일")),
            "work_end_date": normalize_cell(replacement_worker.get("종료일")),
            "employment_duration_days": replacement_duration,
            "is_new_hire": str(replacement_worker.get("신규채용") or "").upper()
            == "Y",
        },
        "source_feature_summary": {
            "data_type_code": features.get("data_type_code"),
            "data_type_label": features.get("data_type_label"),
            "program_type_source": features.get("program_type_source"),
            "judgment_reason_text": features.get("judgment_reason_text"),
            "judgment_basis_text": features.get("judgment_basis_text"),
        },
    }

    payload = {
        "policy_source": "demo_fixture",
        "company": {
            "name": company.get("company_name"),
            "business_registration_number": company.get(
                "business_registration_number"
            ),
            "size": "small",
            "insured_employee_count": company.get("insured_employee_count"),
            "is_priority_support_enterprise": None,
            "is_small_or_medium_company": True,
            "has_replacement_worker": has_replacement_worker,
        },
        "employee": {
            "name": target_employee.get("성명"),
            "department_role": target_employee.get("부서/직무"),
            "employment_type": target_employee.get("고용형태"),
            "hire_date": normalize_cell(target_employee.get("입사일")),
            "leave_type": leave_type,
            "salary_monthly": target_employee.get("급여(원)"),
        },
        "leave_event": {
            "leave_type": leave_type,
            "start_date": start_date,
            "end_date": end_date,
            "requested_months": requested_months(start_date, end_date),
            "duration_days": duration_days,
            "has_replacement_worker": has_replacement_worker,
        },
        "replacement_worker": extracted_fields["replacement_worker"],
        "employer_cost_items": [],
    }

    expected_result = {
        "primary_assertion": {
            "policy_id": target_policy_id,
            "expected_status": expected_status,
            "expected_reason_code": reason_code,
            "expected_reason": features.get("judgment_reason_text"),
        },
        "manual_calculation": {
            "leave_start_date": start_date,
            "leave_end_date": end_date,
            "leave_duration_days": duration_days,
            "replacement_worker_hire_date": extracted_fields[
                "replacement_worker"
            ]["hire_date"],
            "replacement_employment_duration_days": replacement_duration,
        },
        "expected_recommendation_comparison": {
            "target_policy_id": target_policy_id,
            "expected_eligible_policy_ids": (
                [target_policy_id] if expected_status == "eligible" else []
            ),
            "expected_rejected_policy_ids": (
                [target_policy_id] if expected_status == "ineligible" else []
            ),
            "expected_rejected_reason_codes": (
                [reason_code] if expected_status == "ineligible" else []
            ),
            "amount_assertion": {
                "expected_amount": None if expected_status == "eligible" else 0,
                "note": "Draft only. Actual amount comparison is intentionally deferred.",
            },
        },
    }

    case_data = {
        "case_id": case_id,
        "fixture_notice": "TEST FIXTURE ONLY. Draft expected recommendation behavior for a mock company upload case. Do not save to DB or treat as approved policy data.",
        "source_folder": str(folder),
        "source_files": {
            key: str(path.name) for key, path in file_paths.items()
        },
        "file_presence": file_presence,
        "generation_status": {
            "status": "hold" if errors else "draft_created",
            "errors": errors,
            "missing_files": missing_files,
        },
        "extracted_fields": extracted_fields,
        "recommendation_engine_input_draft": payload,
        "field_mapping": {
            "mapped_to_current_demo_api": [
                "company.name",
                "company.business_registration_number",
                "company.size",
                "company.has_replacement_worker",
                "employee.name",
                "employee.leave_type",
                "leave_event.leave_type",
                "leave_event.start_date",
                "leave_event.end_date",
                "leave_event.requested_months",
                "leave_event.duration_days",
                "leave_event.has_replacement_worker",
                "replacement_worker.hire_date",
                "replacement_worker.employment_duration_days",
                "employer_cost_items",
            ],
            "missing_or_not_provided": missing_fields,
        },
        "policy_verification_mapping": build_policy_mapping(
            target_policy_id,
            expected_status,
            reason_code,
        ),
        "expected_result": expected_result,
        "comparison_readiness": {
            "actual_vs_expected_not_run": True,
            "comparability_likelihood": comparable_likelihood,
            "rule_input_expansion_required": RULE_INPUT_EXPANSION_BY_POLICY.get(
                target_policy_id,
                [],
            ),
            "notes": [
                "This file is an expected-result draft only.",
                "Actual-result comparison is intentionally deferred for the 20-case expansion step.",
            ],
        },
    }

    return {
        "case_id": case_id,
        "folder_name": folder.name,
        "fixture_path": fixture_path,
        "fixture_filename": fixture_path.name,
        "case_data": case_data,
        "target_policy_id": target_policy_id,
        "expected_status": expected_status,
        "reason_code": reason_code,
        "missing_fields": missing_fields,
        "generation_success": not errors,
        "hold_reasons": errors,
        "comparability_likelihood": comparable_likelihood,
        "rule_input_expansion_required": RULE_INPUT_EXPANSION_BY_POLICY.get(
            target_policy_id,
            [],
        ),
    }


def write_case_fixture(case):
    fixture_path = case["fixture_path"]
    if fixture_path.name == "04_hana_machine_expected.json" and fixture_path.exists():
        return False

    fixture_path.parent.mkdir(parents=True, exist_ok=True)
    fixture_path.write_text(
        json.dumps(case["case_data"], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return True


def build_inventory(cases):
    policy_counts = Counter(case["target_policy_id"] for case in cases)
    status_counts = Counter(case["expected_status"] for case in cases)
    missing_counter = Counter()
    for case in cases:
        missing_counter.update(case["missing_fields"])

    high_comparability = [
        case["case_id"]
        for case in cases
        if case["comparability_likelihood"] == "high"
    ]
    needs_rule_expansion = [
        {
            "case_id": case["case_id"],
            "target_policy_id": case["target_policy_id"],
            "required_fields": case["rule_input_expansion_required"],
        }
        for case in cases
        if case["rule_input_expansion_required"]
    ]

    return {
        "fixture_notice": "TEST INVENTORY ONLY. Generated from mock company folders for expected-result drafting. No actual recommendation comparison was executed.",
        "source_root": str(SOURCE_ROOT),
        "total_company_folder_count": len(cases),
        "expected_result_generation": {
            "success_count": sum(1 for case in cases if case["generation_success"]),
            "hold_count": sum(1 for case in cases if not case["generation_success"]),
        },
        "policy_case_counts": dict(policy_counts),
        "expected_status_counts": dict(status_counts),
        "missing_field_top5": [
            {"field": field, "count": count}
            for field, count in missing_counter.most_common(5)
        ],
        "actual_vs_expected_high_comparability_cases": high_comparability,
        "rule_input_expansion_required_cases": needs_rule_expansion,
        "cases": [
            {
                "case_id": case["case_id"],
                "folder_name": case["folder_name"],
                "fixture_path": str(case["fixture_path"]),
                "fixture_written": case.get("fixture_written", False),
                "generation_success": case["generation_success"],
                "hold_reasons": case["hold_reasons"],
                "target_policy_id": case["target_policy_id"],
                "expected_status": case["expected_status"],
                "reason_code": case["reason_code"],
                "missing_fields": case["missing_fields"],
                "comparability_likelihood": case["comparability_likelihood"],
                "rule_input_expansion_required": case[
                    "rule_input_expansion_required"
                ],
            }
            for case in cases
        ],
    }


def update_plan(inventory):
    lines = [
        "",
        "## 20-Case Expected Draft Inventory",
        "",
        "Generated by:",
        "",
        "`python scripts\\build_mock_company_expected_cases.py`",
        "",
        "Inventory output:",
        "",
        f"`{OUTPUT_PATH}`",
        "",
        "Scope:",
        "",
        "- Scanned 20 mock company folders.",
        "- Checked `기업특징.txt`, `사업자등록증.pdf`, and `직원명단.xlsx` presence.",
        "- Generated expected-result drafts only; no actual recommendation comparison was run for the new cases.",
        "- Existing `04_hana_machine_expected.json` was preserved.",
        "",
        "Summary:",
        "",
        f"- Total folders: {inventory['total_company_folder_count']}",
        f"- Expected drafts generated or preserved: {inventory['expected_result_generation']['success_count']}",
        f"- Holds: {inventory['expected_result_generation']['hold_count']}",
        f"- Policy counts: `{json.dumps(inventory['policy_case_counts'], ensure_ascii=False)}`",
        f"- Status counts: `{json.dumps(inventory['expected_status_counts'], ensure_ascii=False)}`",
        "",
        "Actual-vs-expected readiness:",
        "",
        f"- High-comparability cases: `{', '.join(inventory['actual_vs_expected_high_comparability_cases'])}`",
        "- Rule-input expansion is still needed for parental leave / working-hour reduction duration checks before full comparison.",
        "",
    ]
    existing = PLAN_PATH.read_text(encoding="utf-8")
    marker = "\n## 20-Case Expected Draft Inventory"
    if marker in existing:
        existing = existing.split(marker, 1)[0].rstrip() + "\n"

    PLAN_PATH.write_text(existing + "\n".join(lines), encoding="utf-8")


def main():
    folders = sorted(path for path in SOURCE_ROOT.iterdir() if path.is_dir())
    cases = [build_case(folder, index + 1) for index, folder in enumerate(folders)]

    for case in cases:
        case["fixture_written"] = write_case_fixture(case)

    inventory = build_inventory(cases)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(inventory, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    update_plan(inventory)

    print(
        json.dumps(
            {
                "status": "PASS",
                "total_company_folder_count": inventory[
                    "total_company_folder_count"
                ],
                "success_count": inventory["expected_result_generation"][
                    "success_count"
                ],
                "hold_count": inventory["expected_result_generation"][
                    "hold_count"
                ],
                "policy_case_counts": inventory["policy_case_counts"],
                "expected_status_counts": inventory["expected_status_counts"],
                "inventory_path": str(OUTPUT_PATH),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
